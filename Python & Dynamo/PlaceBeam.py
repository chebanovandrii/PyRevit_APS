import clr
import openpyxl  # to read Excel file data

# Import Revit Services
clr.AddReference('RevitServices')
from RevitServices.Persistence import DocumentManager

# Import Revit API
clr.AddReference('RevitAPI')
from Autodesk.Revit.DB import *
from Autodesk.Revit.Creation import ItemFactoryBase

# Get the current Revit document
doc = DocumentManager.Instance.CurrentDBDocument
uiapp = DocumentManager.Instance.CurrentUIApplication
app = uiapp.Application

# Input: Excel file path, level, and beam type data
excel_file = IN[0]  # Full path to the Excel file
beam_type_name = IN[1]  # Beam family type (e.g., 'W12x35')
level_name = IN[2]  # Level where beams will be placed

# Open the Excel workbook
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Get the beam type from the Revit document by name
beam_type = None
beam_types = FilteredElementCollector(doc).OfClass(FamilySymbol).ToElements()
for beam in beam_types:
    if beam.Name == beam_type_name:
        beam_type = beam
        break

if not beam_type:
    beam_type = beam_types.FirstElement()

# Get the level from the Revit document by name
level = None
levels = FilteredElementCollector(doc).OfClass(Level).ToElements()
for l in levels:
    if l.Name == level_name:
        level = l
        break

if not level:
	level = levels.FirstElement()

# Create beams from Excel data
transaction = Transaction(doc, "Place Beams from Excel")
transaction.Start()

# Iterate through the rows in Excel and place beams
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
    start_point = XYZ(row[0].value, row[1].value, level.Elevation)  # X, Y, Z from Excel
    end_point = XYZ(row[2].value, row[3].value, level.Elevation)  # X, Y, Z from Excel
    
    # Place beam (line element)
    if not beam_type.IsActive:
        beam_type.Activate()
    
    # Create the beam as a structural element
    line = Line.CreateBound(start_point, end_point)
    beam = doc.Create.NewFamilyInstance(line, beam_type, level, StructuralType.Beam)

transaction.Commit()

# Output: List of created beams
OUT = "Beams have been placed successfully."
